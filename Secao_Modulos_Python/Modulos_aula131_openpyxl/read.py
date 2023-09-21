from pathlib import Path
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

# Carregando um arquivo xlsx
workbook = load_workbook(WORKBOOK_PATH)

# Nome da planilha
sheet_name = 'Minha_Planilha'

# Selecionando a planilha de uso
worksheet: Worksheet = workbook[sheet_name]

for row in worksheet.iter_rows(min_row=2):
    for cell in row:
        print(cell.value, end='\t')
    print()


# salvando a workbook
# workbook.save(WORKBOOK_PATH)
