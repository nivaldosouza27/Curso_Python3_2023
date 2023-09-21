from pathlib import Path
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

ROOT_FOLDER = Path(__file__).parent
WORKBOOK_PATH = ROOT_FOLDER / 'workbook.xlsx'

workbook = Workbook()
# worksheet: Worksheet = workbook.active

# Criando uma nova planilha para receber os dados
sheet_name = 'Minha_Planilha'
workbook.create_sheet(sheet_name, 0)

# Selecionando a planilha de uso
worksheet: Worksheet = workbook[sheet_name]

# removendo a planilha padrão
workbook.remove(workbook['Sheet'])

# Criando os cabeçalhos
worksheet.cell(1, 1, 'Nome')
worksheet.cell(1, 2, 'Idade')
worksheet.cell(1, 3, 'Nota')

# Criando a estrutura de dados que serão inseridos
students = [
    ['João', 14, 5.5],
    ['Maria', 13, 9.7],
    ['Luiz', 15, 8.8],
    ['Alberto', 16, 10]
]

# Inserindo dados com o estrutura de repetição
# for i, student_row in enumerate(students, start=2):
#     for j, student_collum in enumerate(student_row, start=1):
#         worksheet.cell(i, j, student_collum)


# Inserindo dados com append iterando uma lista de dados
for student in students:
    worksheet.append(student)

# salvando a workbook
workbook.save(WORKBOOK_PATH)
